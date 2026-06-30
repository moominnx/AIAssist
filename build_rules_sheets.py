# -*- coding: utf-8 -*-
"""Add regulation (กฎระเบียบ) comparison sheets to UniAssist_Extraction_Evaluation.xlsx
Sheets added: 'Rules Coverage', 'Rules Agreement Detail', 'Rules GT Check List'
"""
import json, re, os
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

BASE = "/Users/nms/AIAssist/data/extracted_rules"
XLSX = "/Users/nms/Downloads/UniAssist_Extraction_Evaluation.xlsx"
PROGS = ["IT", "DSBA", "BIT", "AIT"]
PROG_NAME = {
    "IT": "IT (เทคโนโลยีสารสนเทศ)",
    "DSBA": "DSBA (วิทยาการข้อมูลฯ)",
    "BIT": "BIT (IT ทางธุรกิจ นานาชาติ)",
    "AIT": "AIT (เทคโนโลยี AI)",
}

# ---------- theme ----------
NAVY = "FF15457A"
ORANGE = "FFE8762C"
GRAY = "FF555555"
WHITE = "FFFFFFFF"
GREEN = "FFC6EFCE"
YELLOW = "FFFFEB9C"
RED = "FFFFC7CE"
GREENH = "FFE2EFDA"  # light green section header
HDR_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FONT = Font(bold=True, color=WHITE, size=10)
TITLE_FONT = Font(bold=True, color=NAVY, size=14)
SUB_FONT = Font(color=GRAY, size=10, italic=True)
SECT_FONT = Font(bold=True, color=NAVY, size=11)
THIN = Side(style="thin", color="FFD9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")

TH_DIG = str.maketrans("๐๑๒๓๔๕๖๗๘๙", "0123456789")

def norm(s):
    return (s or "").translate(TH_DIG)

# ---------- load ----------
def load(p, model):
    return json.load(open(f"{BASE}/{p}/{p}_{model}_rules.json", encoding="utf-8"))

ds = {p: load(p, "deepseek") for p in PROGS}
gm = {p: load(p, "gemini") for p in PROGS}
ty = {p: load(p, "typhoon") for p in PROGS}
qw = json.load(open(f"{BASE}/rules_qwen.json", encoding="utf-8"))           # institute-wide (shared)
qw_dept = json.load(open(f"{BASE}/department_rules_qwen.json", encoding="utf-8"))  # per-program

# ---------- 16 canonical categories (Family A) ----------
CATS = [
    ("graduation_requirements", "เกณฑ์การสำเร็จการศึกษา"),
    ("honors_criteria", "เกณฑ์เกียรตินิยม"),
    ("dismissal_criteria", "เกณฑ์พ้นสภาพนักศึกษา"),
    ("probation_rules", "เกณฑ์ภาคทัณฑ์ (probation)"),
    ("grading_system", "ระบบเกรด/การคิดคะแนน"),
    ("registration_rules", "เกณฑ์การลงทะเบียน"),
    ("leave_of_absence_rules", "การลาพักการศึกษา"),
    ("withdrawal_rules", "การลาออก"),
    ("examination_rules", "การสอบ/วัดผล"),
    ("academic_misconduct_rules", "การทุจริตทางวิชาการ"),
    ("student_conduct_rules", "ระเบียบความประพฤติ"),
    ("disciplinary_penalties", "บทลงโทษทางวินัย"),
    ("appeal_rules", "การอุทธรณ์"),
    ("readmission_rules", "การกลับเข้าศึกษา"),
    ("transfer_credit_rules", "การเทียบโอนหน่วยกิต"),
    ("other_regulations", "ระเบียบอื่น ๆ"),
]

# Typhoon rule_type -> canonical category
TY_MAP = {
    "graduation": "graduation_requirements",
    "honors": "honors_criteria",
    "dismissal": "dismissal_criteria",
    "gpa": "grading_system",
    "assessment": "examination_rules",
    "registration": "registration_rules",
    "other": "other_regulations",
    # "prerequisite" -> course-level, ไม่นับเป็นหมวดกฎระเบียบ
}

def nonempty(v):
    if v is None:
        return False
    if isinstance(v, (list, dict)):
        if not v:
            return False
        if isinstance(v, dict):
            return any(nonempty(x) for x in v.values())
        return any(nonempty(x) for x in v) if all(not isinstance(x,(dict,list)) for x in v) else True
    if isinstance(v, str):
        return v.strip() != ""
    return True

def cat_present_A(doc, key):
    return nonempty(doc.get(key))

# Qwen coverage: institute file (shared) for most; graduation also satisfied by dept file
def qwen_present(key):
    return cat_present_A(qw, key)

# Typhoon coverage per program
def ty_cats(p):
    covered = set()
    for x in ty[p]:
        c = TY_MAP.get(x.get("rule_type"))
        if c:
            covered.add(c)
    return covered

# ======================================================================
# SHEET 1: Rules Coverage
# ======================================================================
wb = openpyxl.load_workbook(XLSX)
for nm in ["Rules Coverage", "Rules Agreement Detail", "Rules GT Check List"]:
    if nm in wb.sheetnames:
        del wb[nm]

ws = wb.create_sheet("Rules Coverage")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws["B2"] = "Rules Coverage — หมวดกฎระเบียบที่แต่ละโมเดล extract ได้ (16 หมวด × 4 หลักสูตร)"
ws["B2"].font = TITLE_FONT
ws["B3"] = ("เทียบ 4 โมเดล: Qwen ใช้ข้อบังคับสถาบันชุดเดียวร่วมทุกหลักสูตร | "
            "Typhoon เป็น flat-list map จาก rule_type | ✓ = พบ, ✗ = ไม่พบ")
ws["B3"].font = SUB_FONT

headers = ["หลักสูตร", "หมวดกฎระเบียบ", "Qwen", "Gemini", "DeepSeek", "Typhoon", "พบกี่โมเดล (/4)"]
widths = [30, 30, 10, 10, 11, 10, 16]
r0 = 5
for i, h in enumerate(headers):
    c = ws.cell(r0, 2 + i, h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR; c.border = BORDER
    ws.column_dimensions[chr(ord("B") + i)].width = widths[i]

cov_rows = []  # (prog, cat_th, key, {model:bool}, count)
r = r0 + 1
for p in PROGS:
    tyset = ty_cats(p)
    for key, th in CATS:
        present = {
            "Qwen": qwen_present(key),
            "Gemini": cat_present_A(gm[p], key),
            "DeepSeek": cat_present_A(ds[p], key),
            "Typhoon": key in tyset,
        }
        cnt = sum(present.values())
        cov_rows.append((p, th, key, present, cnt))
        vals = [PROG_NAME[p], th,
                "✓" if present["Qwen"] else "✗",
                "✓" if present["Gemini"] else "✗",
                "✓" if present["DeepSeek"] else "✗",
                "✓" if present["Typhoon"] else "✗",
                f"{cnt}/4"]
        for i, v in enumerate(vals):
            c = ws.cell(r, 2 + i, v)
            c.border = BORDER
            c.alignment = WRAP if i == 1 else CTR
            if i == 0:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
last = r - 1
# conditional formatting on ✓/✗ columns (D..G == cols 4..7 -> B..)
chk_range = f"D{r0+1}:G{last}"
ws.conditional_formatting.add(chk_range,
    FormulaRule(formula=['D6="✓"'], fill=PatternFill("solid", fgColor=GREEN)))
ws.conditional_formatting.add(chk_range,
    FormulaRule(formula=['D6="✗"'], fill=PatternFill("solid", fgColor=RED)))
# count column H (col 8)
cnt_range = f"H{r0+1}:H{last}"
ws.conditional_formatting.add(cnt_range, FormulaRule(formula=['$H6="4/4"'], fill=PatternFill("solid", fgColor=GREEN)))
ws.conditional_formatting.add(cnt_range, FormulaRule(formula=['OR($H6="3/4",$H6="2/4")'], fill=PatternFill("solid", fgColor=YELLOW)))
ws.conditional_formatting.add(cnt_range, FormulaRule(formula=['OR($H6="1/4",$H6="0/4")'], fill=PatternFill("solid", fgColor=RED)))
ws.freeze_panes = "B6"

# ======================================================================
# value extraction for AGREEMENT
# ======================================================================
def f2(x):
    """normalize numeric to string for compare"""
    if x is None:
        return None
    try:
        fx = float(x)
        if fx == int(fx):
            return str(int(fx))
        return f"{fx:g}"
    except (TypeError, ValueError):
        return str(x).strip()

def find_int(text, pat):
    m = re.search(pat, norm(text))
    return m.group(1) if m else None

# ---- minimum credits ----
def mc_ds(p):  return f2(ds[p]["graduation_requirements"].get("minimum_credits"))
def mc_gm(p):  return f2(gm[p]["graduation_requirements"].get("minimum_credits"))
def mc_qw(p):  return f2(qw_dept[p]["graduation_requirements"].get("minimum_credits"))  # per-program from dept file
def mc_ty(p):
    return None  # Typhoon ไม่มีค่าหน่วยกิตจบเป็นตัวเลข

# ---- minimum gpa ----
def gpa_ds(p): return f2(ds[p]["graduation_requirements"].get("minimum_gpa"))
def gpa_gm(p): return f2(gm[p]["graduation_requirements"].get("minimum_gpa"))
def gpa_qw(p): return f2(qw["graduation_requirements"].get("minimum_gpa"))  # institute shared = 2.0
def gpa_ty(p):
    vals = [x.get("value") for x in ty[p] if x.get("rule_type") == "graduation" and x.get("value")]
    return f2(vals[0]) if vals else None

# ---- honors gpa ----
def honor_A(doc, slot):
    h = doc.get("honors_criteria", {})
    return f2(h.get(slot, {}).get("gpa_min")) if isinstance(h, dict) else None
def honor_qw(slot):
    return f2(qw["honors_criteria"].get(slot, {}).get("gpa_min"))
def honor_ty(p, slot):
    """slot in {gold, first, second}; classify Typhoon honors records by description
    so 'อันดับหนึ่ง' inside the gold-medal text doesn't leak into first_class."""
    for x in ty[p]:
        if x.get("rule_type") != "honors" or not x.get("value"):
            continue
        desc = x.get("description") or ""
        if "ทอง" in desc:
            cls = "gold"
        elif "อันดับสอง" in desc:
            cls = "second"
        elif "อันดับหนึ่ง" in desc:
            cls = "first"
        else:
            continue
        if cls == slot:
            return f2(x.get("value"))
    return None

# ---- registration credits (regex on concatenated text) ----
def reg_text(doc_reg):
    if isinstance(doc_reg, dict):
        items = doc_reg.get("rules", [])
    elif isinstance(doc_reg, list):
        items = doc_reg
    else:
        items = []
    return " ".join(items)

# anchored patterns: exception clause uses "...น้อยกว่า X" (not "ไม่น้อยกว่า"),
# normal max uses "และไม่เกิน X", special-case max uses "ทั้งนี้ ต้องไม่เกิน X"
P_MIN = r"ไม่น้อยกว่า\s*(\d+)\s*หน่วยกิต"
P_MAX = r"และไม่เกิน\s*(\d+)\s*หน่วยกิต"
P_SPECIAL = r"ทั้งนี้\s*ต้องไม่เกิน\s*(\d+)"

def reg_min(doc):
    r = doc.get("registration_rules")
    if isinstance(r, dict) and r.get("credits_per_semester_min") is not None:
        return f2(r["credits_per_semester_min"])  # DeepSeek structured (real value, even if wrong)
    return find_int(reg_text(r), P_MIN)
def reg_max(doc):
    r = doc.get("registration_rules")
    if isinstance(r, dict) and r.get("credits_per_semester_max") is not None:
        return f2(r["credits_per_semester_max"])
    return find_int(reg_text(r), P_MAX)
def reg_special(doc):
    return find_int(reg_text(doc.get("registration_rules")), P_SPECIAL)

def reg_qw_text():
    return reg_text(qw.get("registration_rules"))

ATTRS = [
    ("เกณฑ์การสำเร็จการศึกษา", "หน่วยกิตขั้นต่ำที่ต้องเรียนจบ",
     lambda p: mc_qw(p), lambda p: mc_gm(p), lambda p: mc_ds(p), lambda p: mc_ty(p)),
    ("เกณฑ์การสำเร็จการศึกษา", "GPA สะสมขั้นต่ำที่จบได้",
     lambda p: gpa_qw(p), lambda p: gpa_gm(p), lambda p: gpa_ds(p), lambda p: gpa_ty(p)),
    ("เกณฑ์เกียรตินิยม", "GPA เกียรตินิยมอันดับ 1 เหรียญทอง",
     lambda p: honor_qw("first_class_gold_medal"), lambda p: honor_A(gm[p],"first_class_gold_medal"),
     lambda p: honor_A(ds[p],"first_class_gold_medal"), lambda p: honor_ty(p,"gold")),
    ("เกณฑ์เกียรตินิยม", "GPA เกียรตินิยมอันดับ 1",
     lambda p: honor_qw("first_class"), lambda p: honor_A(gm[p],"first_class"),
     lambda p: honor_A(ds[p],"first_class"), lambda p: honor_ty(p,"first")),
    ("เกณฑ์เกียรตินิยม", "GPA เกียรตินิยมอันดับ 2",
     lambda p: honor_qw("second_class"), lambda p: honor_A(gm[p],"second_class"),
     lambda p: honor_A(ds[p],"second_class"), lambda p: honor_ty(p,"second")),
    # registration credit limits = institute-wide; Typhoon flat records ไม่เก็บค่านี้น่าเชื่อถือ -> N/A
    ("เกณฑ์การลงทะเบียน", "หน่วยกิตลงทะเบียนขั้นต่ำ/ภาคปกติ",
     lambda p: find_int(reg_qw_text(), P_MIN),
     lambda p: reg_min(gm[p]), lambda p: reg_min(ds[p]), lambda p: None),
    ("เกณฑ์การลงทะเบียน", "หน่วยกิตลงทะเบียนสูงสุด/ภาคปกติ",
     lambda p: find_int(reg_qw_text(), P_MAX),
     lambda p: reg_max(gm[p]), lambda p: reg_max(ds[p]), lambda p: None),
    ("เกณฑ์การลงทะเบียน", "หน่วยกิตสูงสุดกรณีพิเศษ (เพื่อจบ)",
     lambda p: find_int(reg_qw_text(), P_SPECIAL),
     lambda p: reg_special(gm[p]), lambda p: reg_special(ds[p]), lambda p: None),
]

def agree_result(vals):
    """vals: list of (model, value or None). Returns (text, present_count)."""
    present = [(m, v) for m, v in vals if v not in (None, "")]
    n = len(present)
    if n == 0:
        return "— ไม่มีโมเดลใดให้ค่า", 0
    if n == 1:
        return f"⚪ มีโมเดลเดียว ({present[0][0]})", 1
    c = Counter(v for _, v in present)
    top, topc = c.most_common(1)[0]
    if topc == n:
        return f"✅ ตรงกันหมด ({n}/{n})", n
    if topc >= 2:
        return f"⚠️ ตรงกันบางส่วน ({topc}/{n})", n
    return f"❌ ไม่ตรงกันเลย ({n}/{n}) - เช็คก่อน", n

# ======================================================================
# SHEET 2: Rules Agreement Detail
# ======================================================================
ws2 = wb.create_sheet("Rules Agreement Detail")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 3
ws2["B2"] = "Rules Agreement Detail — เทียบ 'ค่าตัวเลขเกณฑ์' ที่แต่ละโมเดลดึงมา ตรงกันหรือไม่"
ws2["B2"].font = TITLE_FONT
ws2["B3"] = ("flag: ✅ เขียว = ทุกโมเดลตรงกัน, ⚠️ เหลือง = ตรงบางส่วน, ❌ แดง = ไม่ตรงเลย, "
            "⚪ = มีโมเดลเดียวเทียบไม่ได้ | ช่องว่าง = โมเดลไม่ได้ดึงค่านี้")
ws2["B3"].font = SUB_FONT
headers2 = ["หลักสูตร", "หมวด", "Attribute (เกณฑ์)", "Qwen", "Gemini", "DeepSeek", "Typhoon", "ผลเทียบ (Auto)"]
widths2 = [26, 22, 34, 12, 12, 12, 12, 26]
r0 = 5
for i, h in enumerate(headers2):
    c = ws2.cell(r0, 2 + i, h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR; c.border = BORDER
    ws2.column_dimensions[chr(ord("B") + i)].width = widths2[i]

agr_rows = []  # (prog, cat, attr, {model:val}, result_text, present_count)
r = r0 + 1
for p in PROGS:
    for cat, attr, fq, fg, fd, ft in ATTRS:
        vq, vg, vd, vt = fq(p), fg(p), fd(p), ft(p)
        res, ncnt = agree_result([("Qwen", vq), ("Gemini", vg), ("DeepSeek", vd), ("Typhoon", vt)])
        agr_rows.append((p, cat, attr, {"Qwen": vq, "Gemini": vg, "DeepSeek": vd, "Typhoon": vt}, res, ncnt))
        row = [PROG_NAME[p], cat, attr,
               vq if vq is not None else "",
               vg if vg is not None else "",
               vd if vd is not None else "",
               vt if vt is not None else "",
               res]
        for i, v in enumerate(row):
            c = ws2.cell(r, 2 + i, v)
            c.border = BORDER
            c.alignment = WRAP if i in (1, 2, 7) else CTR
        r += 1
last2 = r - 1
res_range = f"I{r0+1}:I{last2}"
ws2.conditional_formatting.add(res_range, FormulaRule(formula=['LEFT(I6,1)="✅"'], fill=PatternFill("solid", fgColor=GREEN)))
ws2.conditional_formatting.add(res_range, FormulaRule(formula=['LEFT(I6,1)="⚠"'], fill=PatternFill("solid", fgColor=YELLOW)))
ws2.conditional_formatting.add(res_range, FormulaRule(formula=['LEFT(I6,1)="❌"'], fill=PatternFill("solid", fgColor=RED)))
ws2.freeze_panes = "B6"

# ======================================================================
# SHEET 3: Rules GT Check List
# ======================================================================
ws3 = wb.create_sheet("Rules GT Check List")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 3
ws3["B2"] = "Rules GT Check List — รายการกฎระเบียบที่ต้องเปิดเล่ม มคอ.2 จริงเพื่อตรวจสอบ"
ws3["B2"].font = TITLE_FONT
ws3["B3"] = ("รวม 2 กลุ่ม: (A) หมวดที่โมเดลเก็บไม่ครบ 4/4  (B) ค่าตัวเลขที่โมเดลไม่ตรงกัน — "
            "กรอกคอลัมน์ขวาเทียบกับเล่มจริง")
ws3["B3"].font = SUB_FONT

headers3 = ["หลักสูตร", "หมวด / Attribute", "ประเด็นที่ต้องตรวจ", "สรุปจากโมเดล",
            "เช็คกับเล่มแล้ว?", "ผลจริงจากเล่ม", "หมายเหตุ"]
widths3 = [26, 30, 34, 34, 14, 26, 26]
r0 = 5
for i, h in enumerate(headers3):
    c = ws3.cell(r0, 2 + i, h)
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR; c.border = BORDER
    ws3.column_dimensions[chr(ord("B") + i)].width = widths3[i]

def section(ws, r, text):
    c = ws.cell(r, 2, text)
    c.font = Font(bold=True, color=NAVY, size=11)
    c.fill = PatternFill("solid", fgColor="FFDDE6F0")
    for cc in range(2, 9):
        ws.cell(r, cc).fill = PatternFill("solid", fgColor="FFDDE6F0")
        ws.cell(r, cc).border = BORDER
    return r + 1

def model_summary(present):
    yes = [m for m in ["Qwen", "Gemini", "DeepSeek", "Typhoon"] if present[m]]
    no = [m for m in ["Qwen", "Gemini", "DeepSeek", "Typhoon"] if not present[m]]
    return f"พบใน: {', '.join(yes) if yes else '—'} | ขาด: {', '.join(no) if no else '—'}"

r = r0 + 1
# Section A: coverage gaps (count < 4)
gaps = [x for x in cov_rows if x[4] < 4]
r = section(ws3, r, f"A) หมวดที่โมเดลเก็บไม่ครบ 4/4 ({len(gaps)} รายการ)")
for p, th, key, present, cnt in gaps:
    row = [PROG_NAME[p], th, f"เล่มจริงมีหมวดนี้หรือไม่ ({cnt}/4 โมเดลเก็บได้)",
           model_summary(present), "", "", ""]
    for i, v in enumerate(row):
        c = ws3.cell(r, 2 + i, v)
        c.border = BORDER
        c.alignment = WRAP
        if i == 4:
            c.alignment = CTR
        if i in (5, 6):
            c.fill = PatternFill("solid", fgColor="FFFFF8E7")  # fill-in highlight
    r += 1

r += 1
# Section B: value disagreements (result not ✅ and at least 2 models gave value, OR single-model)
disagree = [x for x in agr_rows if not x[4].startswith("✅")]
r = section(ws3, r, f"B) ค่าตัวเลขเกณฑ์ที่โมเดลไม่ตรงกัน / ตรวจสอบ ({len(disagree)} รายการ)")
for p, cat, attr, vals, res, ncnt in disagree:
    summ = " | ".join(f"{m}={vals[m]}" for m in ["Qwen","Gemini","DeepSeek","Typhoon"] if vals[m] not in (None,""))
    if not summ:
        summ = "ไม่มีโมเดลใดให้ค่า"
    row = [PROG_NAME[p], f"{cat} › {attr}", res.split(" - ")[0], summ, "", "", ""]
    for i, v in enumerate(row):
        c = ws3.cell(r, 2 + i, v)
        c.border = BORDER
        c.alignment = WRAP
        if i == 4:
            c.alignment = CTR
        if i in (5, 6):
            c.fill = PatternFill("solid", fgColor="FFFFF8E7")
    r += 1
ws3.freeze_panes = "B6"

# reorder: place new sheets right after existing 'Attribute Agreement Detail'
order = wb.sheetnames
for nm in ["Rules Coverage", "Rules Agreement Detail", "Rules GT Check List"]:
    order.remove(nm)
idx = order.index("Attribute Agreement Detail") + 1
new_order = order[:idx] + ["Rules Coverage", "Rules Agreement Detail", "Rules GT Check List"] + order[idx:]
wb._sheets.sort(key=lambda s: new_order.index(s.title))

wb.save(XLSX)

# ---- console report ----
print("SAVED:", XLSX)
print("\n=== Coverage summary (count of models per category, per program) ===")
for p in PROGS:
    line = []
    for x in cov_rows:
        if x[0] == p:
            line.append(f"{x[4]}")
    print(f"{p}: " + " ".join(line))
print("\n=== Coverage gaps (<4) ===", len(gaps))
for p, th, key, present, cnt in gaps:
    print(f"  {p:4} {th:28} {cnt}/4  ขาด:{[m for m in ['Qwen','Gemini','DeepSeek','Typhoon'] if not present[m]]}")
print("\n=== Agreement results ===")
from collections import Counter as C2
rescount = C2(x[4].split(" (")[0].split(" - ")[0] for x in agr_rows)
for k,v in rescount.items(): print(f"  {k}: {v}")
print("\n  Disagreements / single / missing detail:")
for p, cat, attr, vals, res, n in agr_rows:
    if not res.startswith("✅"):
        print(f"  {p:4} {attr:34} {res:28} -> {{Q:{vals['Qwen']}, G:{vals['Gemini']}, D:{vals['DeepSeek']}, T:{vals['Typhoon']}}}")
