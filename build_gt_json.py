import openpyxl, json, os, datetime

SRC = "/Users/nms/Downloads/GT_Template-2.xlsx"
OUT_DIR = "/Users/nms/AIAssist/data/gt"

wb = openpyxl.load_workbook(SRC, data_only=True)

def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return v

def strip_star(s):
    if s is None:
        return None
    s = str(s).replace("\n", " ").strip()
    if s.endswith("*"):
        s = s[:-1].strip()
    return s

def write_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("wrote", path)

COURSE_FIELDS = [
    "code", "name_th", "name_en", "credits", "year", "semester",
    "category", "type", "prerequisite", "flexible_year_semester", "note",
]
GENED_FIELDS = [
    "code", "name_th", "name_en", "credits",
    "category", "type", "prerequisite", "flexible_year_semester", "note",
]

def clean_field(field, v):
    # Excel autocorrects typed "ปี/เทอม" like "4/2" into a date (month=ปี, day=เทอม).
    # Recover the intended "year/semester" text instead of an ISO date string.
    if field == "flexible_year_semester" and isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.month}/{v.day}"
    return clean(v)

def convert_course_sheet(ws, fields):
    courses = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        vals = row[1:1 + len(fields)]
        if not any(v not in (None, "") for v in vals):
            continue
        code = clean(vals[0])
        if code is None:
            continue
        entry = {}
        for f, v in zip(fields, vals):
            entry[f] = clean_field(f, v)
        entry["code"] = str(entry["code"])
        courses.append(entry)
    return courses

# --- Academic Plan GT sheets: one JSON file per sheet, under data/gt/<PROGRAM>/ ---
academic_plan_map = {
    "Academic Plan GT — IT no coop": ("IT", "no_coop"),
    "Academic Plan GT — IT coop": ("IT", "coop"),
    "Academic Plan GT — BIT no coop": ("BIT", "no_coop"),
    "Academic Plan GT — BIT coop": ("BIT", "coop"),
    "Academic Plan GT — DSBA N0 coop": ("DSBA", "no_coop"),
    "Academic Plan GT — DSBA coop": ("DSBA", "coop"),
    "Academic Plan GT — AIT": ("AIT", None),
}

for sheet_name, (program, plan) in academic_plan_map.items():
    ws = wb[sheet_name]
    courses = convert_course_sheet(ws, COURSE_FIELDS)
    plan_label = plan or "plan"
    out = {
        "source": f"GT_Template-2.xlsx / {sheet_name}",
        "description": f"Ground Truth รายวิชาหลักสูตร {program}"
        + (f" (แผน {plan})" if plan else ""),
        "program": program,
        "plan": plan,
        "courses": courses,
    }
    fname = f"{program}_academic_plan_{plan}.json" if plan else f"{program}_academic_plan.json"
    write_json(os.path.join(OUT_DIR, program, fname), out)
    print(f"  {program} ({plan_label}): {len(courses)} courses")

# --- General Education sheet: shared across programs, kept at data/gt/ top level ---
ws = wb["General Education"]
gened = convert_course_sheet(ws, GENED_FIELDS)
out = {
    "source": "GT_Template-2.xlsx / General Education sheet",
    "description": "Ground Truth รายวิชาหมวดศึกษาทั่วไป (ใช้ร่วมกันทุกหลักสูตร)",
    "courses": gened,
}
write_json(os.path.join(OUT_DIR, "general_education_ground_truth.json"), out)
print("  general education:", len(gened), "courses")

# --- Rules GT sheet: split per program into data/gt/<PROGRAM>/ ---
def present(s):
    s = clean(s)
    if s == "มี":
        return True
    if s == "ไม่มี":
        return False
    return None

ws = wb["Rules GT"]
programs = {}
order = []
for row in ws.iter_rows(min_row=6, values_only=True):
    row = list(row) + [None] * (11 - len(row))
    prog = clean(row[1])
    cat = clean(row[2])
    if prog not in ("IT", "DSBA", "BIT", "AIT") or not cat:
        continue
    values = []
    for vi, li in [(4, 5), (6, 7), (8, 9)]:
        v = clean(row[vi])
        lab = strip_star(row[li])
        if v is not None or lab:
            values.append({"value": str(v) if v is not None else None, "label": lab})
    entry = {
        "category": cat,
        "present": present(row[3]),
        "values": values,
        "summary": clean(row[10]),
    }
    if prog not in programs:
        programs[prog] = []
        order.append(prog)
    programs[prog].append(entry)

out = {
    "source": "GT_Template-2.xlsx / Rules GT sheet",
    "description": "Ground Truth กฎระเบียบหลักสูตร 16 หมวด × 4 หลักสูตร (จากเล่ม มคอ.2 จริง)",
    "programs": {p: programs[p] for p in order},
}
write_json(os.path.join(OUT_DIR, "rules_ground_truth.json"), out)
for prog in order:
    print(f"  {prog} rules: {len(programs[prog])} categories")
